"""文档解析：FAQ 模板（xlsx/csv）与普通文本（md/txt/pdf/docx）→ Chunk 记录。

FAQ 模板列结构（04 §4.6 变更）：问题 | 答案 | 分类（可选） | 标签（可选）。
文本类按 500 字/50 字重叠分块（04 §3.5 新增）。
"""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from app.core.config import get_settings
from app.core.exceptions import BadRequestError
from app.pipeline.chunker import DEFAULT_CHUNK_SIZE, DEFAULT_OVERLAP, chunk_text

ALLOWED_EXTENSIONS = {".xlsx", ".csv", ".md", ".txt", ".pdf", ".docx"}
MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB（04 §3.5）

ANSWER_MAX_LENGTH = 2000  # 答案截断上限（08 §4.2 建议 2000 字）
QUESTION_MAX_LENGTH = 200


def get_allowed_extensions() -> set[str]:
    """允许上传的扩展名（08 §8：以配置为准，常量仅作默认值）。"""
    return set(get_settings().allowed_extensions)


def get_max_upload_size() -> int:
    """最大上传字节数（与前端 nginx client_max_body_size 保持一致）。"""
    return get_settings().max_upload_size_mb * 1024 * 1024


def parse_tags(raw: str | None) -> list[str]:
    """标签解析：支持英文逗号/中文逗号/顿号/分号分隔；单个 ≤20 字，最多 10 个。"""
    if not raw:
        return []
    tags: list[str] = []
    for part in re.split(r"[,，;；、]", raw):
        tag = part.strip()[:20]
        if tag and tag not in tags:
            tags.append(tag)
    return tags[:10]


def _clean_row(question: str, answer: str, category: str | None, tags_raw: str | None) -> dict[str, Any]:
    question = (question or "").strip()
    answer = (answer or "").strip()
    if not question or not answer:
        return {}
    if len(question) > QUESTION_MAX_LENGTH:
        question = question[:QUESTION_MAX_LENGTH]
    if len(answer) > ANSWER_MAX_LENGTH:
        answer = answer[:ANSWER_MAX_LENGTH]
    return {
        "question": question,
        "answer": answer,
        "category": (category or "").strip()[:50] or None,
        "tags": parse_tags(tags_raw),
        "page": None,
        "row": None,
    }


def _is_header_row(values: list[str]) -> bool:
    joined = "".join(v or "" for v in values)
    return "问题" in joined and "答案" in joined


def _parse_faq_rows(rows: list[list[str]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row_index, values in enumerate(rows, start=1):
        values = [(v or "").strip() for v in values]
        if not any(values):
            continue
        if row_index == 1 and _is_header_row(values):
            continue
        question = values[0] if len(values) > 0 else ""
        answer = values[1] if len(values) > 1 else ""
        category = values[2] if len(values) > 2 else None
        tags_raw = values[3] if len(values) > 3 else None
        record = _clean_row(question, answer, category, tags_raw)
        if record:
            record["row"] = str(row_index)
            records.append(record)
    return records


def _parse_xlsx(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = [[cell if cell is not None else "" for cell in row] for row in sheet.iter_rows(values_only=True)]
    return _parse_faq_rows(rows)


def _parse_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    return _parse_faq_rows(rows)


def _text_chunks(
    text: str,
    page: str | None = None,
    chunk_size: int = DEFAULT_CHUNK_SIZE,
    overlap: int = DEFAULT_OVERLAP,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for segment in chunk_text(text, chunk_size=chunk_size, overlap=overlap):
        question = segment[:50]
        records.append(
            {
                "question": question,
                "answer": segment,
                "category": None,
                "tags": [],
                "page": page,
                "row": None,
            }
        )
    return records


def _parse_txt_md(path: Path, chunk_size: int, overlap: int) -> list[dict[str, Any]]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    return _text_chunks(text, chunk_size=chunk_size, overlap=overlap)


def _parse_pdf(path: Path, chunk_size: int, overlap: int) -> list[dict[str, Any]]:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    records: list[dict[str, Any]] = []
    for page_index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        records.extend(
            _text_chunks(text, page=str(page_index), chunk_size=chunk_size, overlap=overlap)
        )
    return records


def _parse_docx(path: Path, chunk_size: int, overlap: int) -> list[dict[str, Any]]:
    from docx import Document as DocxDocument

    doc = DocxDocument(str(path))
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    return _text_chunks(text, chunk_size=chunk_size, overlap=overlap)


def parse_document(
    path: str | Path,
    file_type: str,
    chunk_size: int = DEFAULT_CHUNK_SIZE,
    overlap: int = DEFAULT_OVERLAP,
) -> list[dict[str, Any]]:
    """解析文件为 Chunk 记录列表（问题/答案/分类/标签/页码/行号）。"""
    path = Path(path)
    ext = f".{file_type.lower().lstrip('.')}" if not file_type.startswith(".") else file_type.lower()
    if ext not in get_allowed_extensions():
        raise BadRequestError(f"不支持的文件类型：{ext}")

    if ext == ".xlsx":
        records = _parse_xlsx(path)
    elif ext == ".csv":
        records = _parse_csv(path)
    elif ext in (".md", ".txt"):
        records = _parse_txt_md(path, chunk_size, overlap)
    elif ext == ".pdf":
        records = _parse_pdf(path, chunk_size, overlap)
    elif ext == ".docx":
        records = _parse_docx(path, chunk_size, overlap)
    else:
        records = []

    if not records:
        raise BadRequestError("未解析到有效内容，请检查文件格式")
    return records
